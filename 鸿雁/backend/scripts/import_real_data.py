#!/usr/bin/env python3
"""导入真实案例库 + 山大人才库到数据库（幂等，可重复执行）

用法: python scripts/import_real_data.py   （在 backend/ 下执行）
数据: ../data/ 下两份 xlsx
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

import openpyxl
from app.database import SessionLocal, engine, Base
from app.models import BorderDemand, MainlandSupply, CompletedAchievement, SduTalent

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'data')
CASE_XLSX = os.path.join(DATA_DIR, '东西协作成果-边疆需求-内地供给案例库（整合版）.xlsx')
TALENT_XLSX = os.path.join(DATA_DIR, '山东大学科研与人才资源.xlsx')


def s(v) -> str:
    return '' if v is None else str(v).strip()


def main():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    wb = openpyxl.load_workbook(CASE_XLSX, data_only=True)

    ws = wb['已完成成果']
    db.query(CompletedAchievement).delete()
    n_ach = 0
    for r in range(4, ws.max_row + 1):
        if not s(ws.cell(r, 1).value):
            continue
        db.add(CompletedAchievement(
            achievement_id=s(ws.cell(r, 1).value), title=s(ws.cell(r, 2).value),
            region=s(ws.cell(r, 3).value), parties=s(ws.cell(r, 4).value),
            finish_time=s(ws.cell(r, 5).value), work_done=s(ws.cell(r, 6).value),
            highlights=s(ws.cell(r, 7).value), achievement_type=s(ws.cell(r, 8).value),
            replicable_points=s(ws.cell(r, 9).value), image_note=s(ws.cell(r, 10).value),
            image_link=s(ws.cell(r, 11).value), evidence=s(ws.cell(r, 12).value),
            source_level=s(ws.cell(r, 13).value), source_body=s(ws.cell(r, 14).value),
            publish_date=s(ws.cell(r, 15).value), source_url=s(ws.cell(r, 16).value),
            verification_status=s(ws.cell(r, 17).value), boundary_note=s(ws.cell(r, 18).value)))
        n_ach += 1

    ws = wb['边疆需求侧']
    db.query(BorderDemand).delete()
    n_dem = 0
    for r in range(4, ws.max_row + 1):
        if not s(ws.cell(r, 1).value):
            continue
        db.add(BorderDemand(
            demand_id=s(ws.cell(r, 1).value), title=s(ws.cell(r, 2).value),
            province=s(ws.cell(r, 3).value), coverage=s(ws.cell(r, 4).value),
            location_detail=s(ws.cell(r, 5).value), publisher=s(ws.cell(r, 6).value),
            pain_point=s(ws.cell(r, 7).value), description=s(ws.cell(r, 8).value),
            expected_goal=s(ws.cell(r, 9).value), stage=s(ws.cell(r, 10).value),
            supply_tags=s(ws.cell(r, 11).value), contact=s(ws.cell(r, 12).value),
            source_level=s(ws.cell(r, 13).value), publish_date=s(ws.cell(r, 14).value),
            source_url=s(ws.cell(r, 15).value),
            verification_status=s(ws.cell(r, 17).value)))
        n_dem += 1

    ws = wb['内地供给侧']
    db.query(MainlandSupply).delete()
    n_sup = 0
    for r in range(4, ws.max_row + 1):
        if not s(ws.cell(r, 1).value):
            continue
        db.add(MainlandSupply(
            supply_id=s(ws.cell(r, 1).value), provider=s(ws.cell(r, 2).value),
            location=s(ws.cell(r, 3).value), subject_type=s(ws.cell(r, 4).value),
            services=s(ws.cell(r, 5).value), tech_advantages=s(ws.cell(r, 6).value),
            use_cases=s(ws.cell(r, 7).value), border_fit=s(ws.cell(r, 8).value),
            delivery_mode=s(ws.cell(r, 9).value), contact=s(ws.cell(r, 10).value),
            source_level=s(ws.cell(r, 12).value) + ' ' + s(ws.cell(r, 11).value),
            publish_date=s(ws.cell(r, 13).value), source_url=s(ws.cell(r, 14).value),
            verification_status=s(ws.cell(r, 15).value)))
        n_sup += 1
    db.commit()

    tb = openpyxl.load_workbook(TALENT_XLSX, data_only=True)
    db.query(SduTalent).delete()
    n_tal = 0
    sheet_domain = {'农业种植与加工': '农业', '能源与光伏': '能源',
                    '产业升级与智能制造': '智能制造', 'AI训练与创新成果': 'AI'}
    for name, domain in sheet_domain.items():
        ws = tb[name]
        hdr = [s(c.value) for c in ws[3]]
        def pick(row, *keys):
            for i, h in enumerate(hdr):
                if h and any(k in h for k in keys):
                    return s(ws.cell(row, i + 1).value)
            return ''
        for r in range(4, ws.max_row + 1):
            org, team = s(ws.cell(r, 1).value), s(ws.cell(r, 3).value)
            if not (org or team):
                continue
            db.add(SduTalent(
                domain=domain, org=org,
                field=pick(r, '研究领域'), team=team or org,
                leader=pick(r, '负责人', '核心教师'), leader_title=pick(r, '人才头衔'),
                patents=pick(r, '专利'), core_tech=pick(r, '核心技术成果'),
                awards=pick(r, '获奖', '项目级别'),
                west_scene=pick(r, '西部场景'), application=pick(r, '具体应用方向'),
                maturity=pick(r, '技术成熟度'), cases=pick(r, '已转化', '合作案例'),
                source_url=pick(r, '官网', '新闻链接'),
                source_note=pick(r, '来源说明') or pick(r, '来源')))
            n_tal += 1
    db.commit()

    print(f'导入完成：已完成成果 {n_ach} | 边疆需求 {n_dem} | 内地供给 {n_sup} | 山大人才 {n_tal}')
    db.close()


if __name__ == '__main__':
    main()
